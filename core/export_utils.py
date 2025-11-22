"""
Export utilities for the Floor Management System.

Provides unified export functionality across all modules supporting:
- CSV export
- Excel export (XLSX)
- PDF export
- Custom column selection
- Filtered data export
"""

import csv
import io
from datetime import datetime
from django.http import HttpResponse
from django.utils.text import slugify


class DataExporter:
    """
    Universal data exporter supporting multiple formats.

    Usage:
        exporter = DataExporter(
            queryset=MyModel.objects.all(),
            fields=['id', 'name', 'status'],
            headers=['ID', 'Name', 'Status'],
            filename='my_export'
        )

        # Export to CSV
        response = exporter.to_csv()

        # Export to Excel
        response = exporter.to_excel()

        # Export to PDF
        response = exporter.to_pdf(title="My Report")
    """

    def __init__(self, queryset, fields, headers=None, filename='export'):
        """
        Initialize exporter.

        Args:
            queryset: Django queryset to export
            fields: List of field names to export
            headers: List of header labels (defaults to fields if not provided)
            filename: Base filename for export (without extension)
        """
        self.queryset = queryset
        self.fields = fields
        self.headers = headers or fields
        self.filename = slugify(filename)

    def _get_field_value(self, obj, field):
        """Get field value from object, handling nested fields and callables."""
        try:
            # Handle nested fields (e.g., 'person__first_name')
            if '__' in field:
                parts = field.split('__')
                value = obj
                for part in parts:
                    value = getattr(value, part, '')
                    if value is None:
                        return ''
                return str(value)

            # Get attribute
            value = getattr(obj, field, '')

            # Handle callable methods
            if callable(value):
                value = value()

            # Handle None
            if value is None:
                return ''

            # Convert to string
            return str(value)
        except Exception:
            return ''

    def _prepare_data(self):
        """Prepare data rows for export."""
        rows = []
        for obj in self.queryset:
            row = []
            for field in self.fields:
                value = self._get_field_value(obj, field)
                row.append(value)
            rows.append(row)
        return rows

    def to_csv(self):
        """Export to CSV format."""
        # Create CSV response
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{self.filename}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'

        # Write CSV
        writer = csv.writer(response)
        writer.writerow(self.headers)

        data = self._prepare_data()
        for row in data:
            writer.writerow(row)

        return response

    def to_excel(self):
        """Export to Excel (XLSX) format."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            # Fallback to CSV if openpyxl not installed
            return self.to_csv()

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Export'

        # Style for headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Write headers
        for col, header in enumerate(self.headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Write data
        data = self._prepare_data()
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # Create response
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{self.filename}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'

        return response

    def to_pdf(self, title='Export Report', orientation='portrait'):
        """
        Export to PDF format.

        Args:
            title: Report title
            orientation: 'portrait' or 'landscape'
        """
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import letter, landscape
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        except ImportError:
            # Fallback to CSV if reportlab not installed
            return self.to_csv()

        # Create PDF buffer
        buffer = io.BytesIO()

        # Set page size
        pagesize = landscape(letter) if orientation == 'landscape' else letter
        doc = SimpleDocTemplate(buffer, pagesize=pagesize)

        # Container for elements
        elements = []
        styles = getSampleStyleSheet()

        # Add title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#4F81BD'),
            spaceAfter=30,
            alignment=1  # Center
        )
        elements.append(Paragraph(title, title_style))
        elements.append(Spacer(1, 0.2 * inch))

        # Add metadata
        meta_style = styles['Normal']
        meta_text = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}<br/>Records: {self.queryset.count()}"
        elements.append(Paragraph(meta_text, meta_style))
        elements.append(Spacer(1, 0.3 * inch))

        # Prepare table data
        data = [self.headers]
        data.extend(self._prepare_data())

        # Limit data if too many rows (PDF performance)
        if len(data) > 1001:  # 1000 data rows + 1 header
            data = data[:1001]
            elements.append(Paragraph(
                f"<b>Note:</b> Only first 1000 records shown. Total: {self.queryset.count()}",
                styles['Normal']
            ))
            elements.append(Spacer(1, 0.2 * inch))

        # Create table
        table = Table(data)

        # Style table
        table.setStyle(TableStyle([
            # Header row
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F81BD')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),

            # Data rows
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F0F0F0')]),
        ]))

        elements.append(table)

        # Build PDF
        doc.build(elements)

        # Get PDF from buffer
        pdf = buffer.getvalue()
        buffer.close()

        # Create response
        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{self.filename}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'

        return response


class ExportHistory:
    """Track user export history."""

    @staticmethod
    def add_export(user, export_type, model_name, record_count):
        """
        Record an export action.

        Args:
            user: User who performed export
            export_type: 'csv', 'excel', or 'pdf'
            model_name: Name of model exported
            record_count: Number of records exported
        """
        from core.models import UserPreference
        from django.utils import timezone

        try:
            pref = UserPreference.get_or_create_for_user(user)

            # Get export history
            export_history = pref.preferences_json.get('export_history', [])

            # Add new export
            export_item = {
                'type': export_type,
                'model': model_name,
                'count': record_count,
                'timestamp': str(timezone.now())
            }

            # Add to beginning
            export_history.insert(0, export_item)

            # Keep only last 50 exports
            export_history = export_history[:50]

            # Save
            pref.preferences_json['export_history'] = export_history
            pref.save()
        except Exception as e:
            # Don't fail the export if history tracking fails
            print(f"Failed to track export history: {e}")

    @staticmethod
    def get_recent_exports(user, limit=10):
        """Get user's recent exports."""
        from core.models import UserPreference

        try:
            pref = UserPreference.get_or_create_for_user(user)
            export_history = pref.preferences_json.get('export_history', [])
            return export_history[:limit]
        except Exception:
            return []


def export_queryset(request, queryset, fields, headers=None, filename='export', format='csv'):
    """
    Helper function to quickly export a queryset.

    Args:
        request: HttpRequest object
        queryset: Django queryset to export
        fields: List of field names
        headers: List of header labels
        filename: Base filename
        format: 'csv', 'excel', or 'pdf'

    Returns:
        HttpResponse with exported file
    """
    exporter = DataExporter(
        queryset=queryset,
        fields=fields,
        headers=headers,
        filename=filename
    )

    # Track export
    if hasattr(request, 'user') and request.user.is_authenticated:
        ExportHistory.add_export(
            user=request.user,
            export_type=format,
            model_name=queryset.model.__name__,
            record_count=queryset.count()
        )

    # Export based on format
    if format == 'excel':
        return exporter.to_excel()
    elif format == 'pdf':
        return exporter.to_pdf(title=f"{queryset.model._meta.verbose_name_plural} Report")
    else:  # csv
        return exporter.to_csv()
