"""
Bulk Import/Export and Batch Operations for Inventory
"""
from django.db import transaction
from decimal import Decimal
import csv
import io
from openpyxl import load_workbook
from .models import (
    Item, ItemCategory, UnitOfMeasure, Supplier,
    StockLevel, StockTransaction, Location, ConditionType, OwnershipType
)


class BulkItemImporter:
    """
    Import items in bulk from Excel or CSV files.
    Expected columns: item_code, name, category_code, uom_code, item_type,
                     reorder_level, description, supplier_code (optional)
    """

    def __init__(self, user):
        self.user = user
        self.errors = []
        self.success_count = 0
        self.update_count = 0

    def import_from_excel(self, file):
        """Import items from Excel file."""
        try:
            wb = load_workbook(file, read_only=True)
            ws = wb.active

            # Skip header row
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            return self._process_rows(rows)
        except Exception as e:
            self.errors.append(f"Excel read error: {str(e)}")
            return False

    def import_from_csv(self, file):
        """Import items from CSV file."""
        try:
            # Decode file content
            text = file.read().decode('utf-8')
            reader = csv.reader(io.StringIO(text))

            # Skip header
            next(reader)
            rows = list(reader)
            return self._process_rows(rows)
        except Exception as e:
            self.errors.append(f"CSV read error: {str(e)}")
            return False

    def _process_rows(self, rows):
        """Process rows from either Excel or CSV."""
        with transaction.atomic():
            for idx, row in enumerate(rows, start=2):
                try:
                    if not row or not row[0]:  # Skip empty rows
                        continue

                    item_code = str(row[0]).strip()
                    name = str(row[1]).strip()
                    category_code = str(row[2]).strip()
                    uom_code = str(row[3]).strip()
                    item_type = str(row[4]).strip().upper()
                    reorder_level = Decimal(row[5]) if row[5] else Decimal('0')
                    description = str(row[6]).strip() if len(row) > 6 and row[6] else ''
                    supplier_code = str(row[7]).strip() if len(row) > 7 and row[7] else None

                    # Validate category
                    try:
                        category = ItemCategory.objects.get(code=category_code)
                    except ItemCategory.DoesNotExist:
                        self.errors.append(f"Row {idx}: Category '{category_code}' not found")
                        continue

                    # Validate UOM
                    try:
                        uom = UnitOfMeasure.objects.get(code=uom_code)
                    except UnitOfMeasure.DoesNotExist:
                        self.errors.append(f"Row {idx}: UOM '{uom_code}' not found")
                        continue

                    # Validate item type
                    valid_types = [choice[0] for choice in Item.ITEM_TYPE_CHOICES]
                    if item_type not in valid_types:
                        self.errors.append(f"Row {idx}: Invalid item type '{item_type}'")
                        continue

                    # Get supplier if provided
                    supplier = None
                    if supplier_code:
                        try:
                            supplier = Supplier.objects.get(code=supplier_code)
                        except Supplier.DoesNotExist:
                            self.errors.append(f"Row {idx}: Supplier '{supplier_code}' not found (item will be created without supplier)")

                    # Create or update item
                    item, created = Item.objects.update_or_create(
                        item_code=item_code,
                        defaults={
                            'name': name,
                            'category': category,
                            'unit_of_measure': uom,
                            'item_type': item_type,
                            'reorder_level': reorder_level,
                            'description': description,
                            'preferred_supplier': supplier,
                            'active': True
                        }
                    )

                    if created:
                        self.success_count += 1
                    else:
                        self.update_count += 1

                except Exception as e:
                    self.errors.append(f"Row {idx}: {str(e)}")

        return len(self.errors) == 0

    def get_summary(self):
        """Get import summary."""
        return {
            'success_count': self.success_count,
            'update_count': self.update_count,
            'error_count': len(self.errors),
            'errors': self.errors
        }


class BulkStockAdjustment:
    """
    Perform bulk stock adjustments from Excel or CSV.
    Expected columns: item_code, location_code, condition_code, ownership_code,
                     quantity, reference, notes
    """

    def __init__(self, user):
        self.user = user
        self.errors = []
        self.success_count = 0

    def process_from_excel(self, file):
        """Process bulk adjustments from Excel."""
        try:
            wb = load_workbook(file, read_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            return self._process_adjustments(rows)
        except Exception as e:
            self.errors.append(f"Excel read error: {str(e)}")
            return False

    def process_from_csv(self, file):
        """Process bulk adjustments from CSV."""
        try:
            text = file.read().decode('utf-8')
            reader = csv.reader(io.StringIO(text))
            next(reader)
            rows = list(reader)
            return self._process_adjustments(rows)
        except Exception as e:
            self.errors.append(f"CSV read error: {str(e)}")
            return False

    def _process_adjustments(self, rows):
        """Process adjustment rows."""
        with transaction.atomic():
            for idx, row in enumerate(rows, start=2):
                try:
                    if not row or not row[0]:
                        continue

                    item_code = str(row[0]).strip()
                    location_code = str(row[1]).strip()
                    condition_code = str(row[2]).strip()
                    ownership_code = str(row[3]).strip()
                    quantity = Decimal(row[4])
                    reference = str(row[5]).strip() if len(row) > 5 and row[5] else f'BULK-ADJ-{idx}'
                    notes = str(row[6]).strip() if len(row) > 6 and row[6] else 'Bulk adjustment'

                    # Get related objects
                    try:
                        item = Item.objects.get(item_code=item_code)
                        location = Location.objects.get(code=location_code)
                        condition = ConditionType.objects.get(code=condition_code)
                        ownership = OwnershipType.objects.get(code=ownership_code)
                    except Item.DoesNotExist:
                        self.errors.append(f"Row {idx}: Item '{item_code}' not found")
                        continue
                    except Location.DoesNotExist:
                        self.errors.append(f"Row {idx}: Location '{location_code}' not found")
                        continue
                    except ConditionType.DoesNotExist:
                        self.errors.append(f"Row {idx}: Condition '{condition_code}' not found")
                        continue
                    except OwnershipType.DoesNotExist:
                        self.errors.append(f"Row {idx}: Ownership '{ownership_code}' not found")
                        continue

                    # Create stock transaction
                    transaction_obj = StockTransaction.objects.create(
                        item=item,
                        to_location=location,
                        condition_type=condition,
                        ownership_type=ownership,
                        transaction_type='ADJUSTMENT',
                        quantity=quantity,
                        reference=reference,
                        notes=notes,
                        performed_by=self.user
                    )

                    # Update stock level
                    stock_level, created = StockLevel.objects.get_or_create(
                        item=item,
                        location=location,
                        condition_type=condition,
                        ownership_type=ownership,
                        defaults={'quantity': Decimal('0')}
                    )
                    stock_level.quantity += quantity
                    stock_level.save()

                    self.success_count += 1

                except Exception as e:
                    self.errors.append(f"Row {idx}: {str(e)}")

        return len(self.errors) == 0

    def get_summary(self):
        """Get adjustment summary."""
        return {
            'success_count': self.success_count,
            'error_count': len(self.errors),
            'errors': self.errors
        }


class BatchItemUpdater:
    """
    Update multiple items at once (e.g., change category, update reorder levels).
    """

    @staticmethod
    def update_category(item_ids, new_category):
        """Update category for multiple items."""
        return Item.objects.filter(id__in=item_ids).update(category=new_category)

    @staticmethod
    def update_reorder_level(item_ids, new_level):
        """Update reorder level for multiple items."""
        return Item.objects.filter(id__in=item_ids).update(reorder_level=new_level)

    @staticmethod
    def activate_items(item_ids):
        """Activate multiple items."""
        return Item.objects.filter(id__in=item_ids).update(active=True)

    @staticmethod
    def deactivate_items(item_ids):
        """Deactivate multiple items."""
        return Item.objects.filter(id__in=item_ids).update(active=False)

    @staticmethod
    def update_supplier(item_ids, new_supplier):
        """Update preferred supplier for multiple items."""
        return Item.objects.filter(id__in=item_ids).update(preferred_supplier=new_supplier)
