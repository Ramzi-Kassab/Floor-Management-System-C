import csv
from io import BytesIO
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime


def export_to_csv(queryset, filename, fields):
    """Export queryset to CSV."""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    writer = csv.writer(response)

    # Write headers
    headers = [field['name'] for field in fields]
    writer.writerow(headers)

    # Write data
    for obj in queryset:
        row = []
        for field in fields:
            value = field['value'](obj) if callable(field['value']) else getattr(obj, field['value'], '')
            row.append(str(value) if value is not None else '')
        writer.writerow(row)

    return response


def export_to_excel(queryset, filename, fields, sheet_name='Data'):
    """Export queryset to Excel with formatting."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Header styling
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    header_alignment = Alignment(horizontal='center', vertical='center')

    # Write headers
    headers = [field['name'] for field in fields]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Write data
    for row_num, obj in enumerate(queryset, 2):
        for col_num, field in enumerate(fields, 1):
            value = field['value'](obj) if callable(field['value']) else getattr(obj, field['value'], '')
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value

    # Auto-adjust column widths
    for col_num, field in enumerate(fields, 1):
        column_letter = get_column_letter(col_num)
        max_length = len(field['name'])
        for cell in ws[column_letter]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Create response
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    return response


def get_stock_export_fields():
    """Define fields for stock level export."""
    return [
        {'name': 'Item Code', 'value': lambda obj: obj.item.item_code},
        {'name': 'Item Name', 'value': lambda obj: obj.item.name},
        {'name': 'Category', 'value': lambda obj: obj.item.category.name},
        {'name': 'Warehouse', 'value': lambda obj: obj.location.warehouse.code},
        {'name': 'Location', 'value': lambda obj: obj.location.code},
        {'name': 'Condition', 'value': lambda obj: obj.condition_type.code},
        {'name': 'Ownership', 'value': lambda obj: obj.ownership_type.code},
        {'name': 'Quantity', 'value': 'quantity'},
        {'name': 'UOM', 'value': lambda obj: obj.item.unit_of_measure.code},
        {'name': 'Last Updated', 'value': lambda obj: obj.last_updated.strftime('%Y-%m-%d %H:%M')},
    ]


def get_transaction_export_fields():
    """Define fields for transaction export."""
    return [
        {'name': 'Date/Time', 'value': lambda obj: obj.performed_at.strftime('%Y-%m-%d %H:%M')},
        {'name': 'Type', 'value': lambda obj: obj.get_transaction_type_display()},
        {'name': 'Item Code', 'value': lambda obj: obj.item.item_code},
        {'name': 'Item Name', 'value': lambda obj: obj.item.name},
        {'name': 'Quantity', 'value': 'quantity'},
        {'name': 'UOM', 'value': lambda obj: obj.item.unit_of_measure.code},
        {'name': 'From Location', 'value': lambda obj: obj.from_location.code if obj.from_location else '-'},
        {'name': 'To Location', 'value': lambda obj: obj.to_location.code if obj.to_location else '-'},
        {'name': 'Condition', 'value': lambda obj: obj.condition_type.code},
        {'name': 'Ownership', 'value': lambda obj: obj.ownership_type.code},
        {'name': 'Reference', 'value': 'reference'},
        {'name': 'Performed By', 'value': lambda obj: obj.performed_by.username},
        {'name': 'Notes', 'value': 'notes'},
    ]


def get_items_export_fields():
    """Define fields for items export."""
    return [
        {'name': 'Item Code', 'value': 'item_code'},
        {'name': 'Name', 'value': 'name'},
        {'name': 'Category', 'value': lambda obj: obj.category.name},
        {'name': 'Type', 'value': lambda obj: obj.get_item_type_display()},
        {'name': 'UOM', 'value': lambda obj: obj.unit_of_measure.code},
        {'name': 'Reorder Level', 'value': 'reorder_level'},
        {'name': 'Preferred Supplier', 'value': lambda obj: obj.preferred_supplier.name if obj.preferred_supplier else '-'},
        {'name': 'Active', 'value': lambda obj: 'Yes' if obj.active else 'No'},
        {'name': 'Description', 'value': 'description'},
    ]


def get_low_stock_export_fields():
    """Define fields for low stock items export."""
    return [
        {'name': 'Item Code', 'value': 'item_code'},
        {'name': 'Name', 'value': 'name'},
        {'name': 'Category', 'value': lambda obj: obj.category.name},
        {'name': 'Current Stock', 'value': lambda obj: obj.total_stock if hasattr(obj, 'total_stock') else 0},
        {'name': 'Reorder Level', 'value': 'reorder_level'},
        {'name': 'UOM', 'value': lambda obj: obj.unit_of_measure.code},
        {'name': 'Stock %', 'value': lambda obj: f"{(obj.total_stock / obj.reorder_level * 100):.0f}%" if hasattr(obj, 'total_stock') and obj.total_stock and obj.reorder_level > 0 else '0%'},
        {'name': 'Preferred Supplier', 'value': lambda obj: obj.preferred_supplier.name if obj.preferred_supplier else '-'},
    ]
